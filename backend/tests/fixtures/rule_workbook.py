from pathlib import Path

from openpyxl import Workbook

from backend.app.rules.excel_importer import JD_CODE_BY_SHEET, SHEET_LAYOUT


def build_rule_workbook(path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in JD_CODE_BY_SHEET:
        worksheet = workbook.create_sheet(sheet_name)
        layout = SHEET_LAYOUT[sheet_name]
        dimension_row = layout.data_start_row
        worksheet.cell(dimension_row, 2, "学历")
        worksheet.cell(dimension_row, 3, 100)
        for index, column in enumerate(layout.tier_cols):
            worksheet.cell(dimension_row, column + 1, index * 10)
        worksheet.cell(dimension_row, layout.keyword_col + 1, "本科、专升本、大专")
        total_row = dimension_row + 1
        worksheet.cell(total_row, 2, "合计总分")
        ranges = ("0-39", "40-69", "70-100", "70-84", "85-100")
        for column, score_range in zip(layout.tier_cols, ranges, strict=False):
            worksheet.cell(total_row, column + 1, score_range)
        if sheet_name == "业务岗全维度评分表格":
            worksheet.cell(total_row + 1, 1, "年龄超过45岁直接淘汰")
    workbook.save(path)
    return path
