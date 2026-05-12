from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl

from backend.app.rules.schema import (
    GradeThreshold,
    HardFilter,
    JudgeDimension,
    RuleDimension,
    RuleSchema,
    Tier,
)


@dataclass(frozen=True)
class SheetLayout:
    data_start_row: int  # first row of dimension data (1-based)
    row_stride: int  # rows per dimension (2 = score row + standard row, 1 = single row)
    tier_cols: tuple[int, ...]  # 0-based column indexes for tier score cells
    tier_labels: tuple[str, ...]  # one label per tier_cols entry; values in {low, mid, high}
    keyword_col: int  # 0-based column index for the keyword cell


# Layout A: 业务岗全维度评分表格. 10 cols, header rows 1+2, dimensions from row 3 in 2-row stride
LAYOUT_A = SheetLayout(
    data_start_row=3,
    row_stride=2,
    tier_cols=(4, 5, 6, 7, 8),
    tier_labels=("low", "low", "mid", "high", "high"),
    keyword_col=9,
)

# Layout B: the other 5 sheets. 8 cols, header row 1, dimensions from row 2 in 1-row stride
LAYOUT_B = SheetLayout(
    data_start_row=2,
    row_stride=1,
    tier_cols=(4, 5, 6),
    tier_labels=("low", "mid", "high"),
    keyword_col=7,
)


JD_CODE_BY_SHEET: dict[str, str] = {
    "业务岗全维度评分表格": "FOREIGN_TRADE",
    "物流代表": "LOGISTICS",
    "采购、产品": "SOURCING_PRODUCT",
    "QC": "QC",
    "SQE": "SQE",
    "项目工程师": "OEM_PROJECT",
}

SHEET_LAYOUT: dict[str, SheetLayout] = {
    "业务岗全维度评分表格": LAYOUT_A,
    "物流代表": LAYOUT_B,
    "采购、产品": LAYOUT_B,
    "QC": LAYOUT_B,
    "SQE": LAYOUT_B,
    "项目工程师": LAYOUT_B,
}

JUDGE_DIM_KEYWORDS = ("独立处理", "情绪稳定", "抗压", "团队", "责任心")

_SPLITTER = re.compile(r"[、,，/\s]+")
_SCORE_NUM = re.compile(r"(\d+(?:\.\d+)?)")
_SCORE_WITH_UNIT = re.compile(r"(\d+(?:\.\d+)?)\s*分")


def _split_keywords(text: str | None) -> list[str]:
    if not text:
        return []
    parts = [p.strip() for p in _SPLITTER.split(text) if p.strip()]
    return parts


def _parse_score(cell: object) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell)
    hits = _SCORE_WITH_UNIT.findall(s)
    if hits:
        return float(hits[-1])
    m = _SCORE_NUM.search(s)
    return float(m.group(1)) if m else None


def _is_judge_dimension(name: str) -> bool:
    return any(k in name for k in JUDGE_DIM_KEYWORDS)


def _parse_hard_filters(notes_cell: str | None) -> list[HardFilter]:
    out: list[HardFilter] = []
    if not notes_cell:
        return out
    text = str(notes_cell)
    if "年龄" in text and "45" in text:
        out.append(
            HardFilter(id="age_max", rule="age <= 45", action="reject", audit_tag="AGE")
        )
    m = re.search(r"总分\s*[＜<]\s*(\d+)", text)
    if m:
        out.append(
            HardFilter(
                id="min_total",
                rule=f"total_score >= {m.group(1)}",
                action="reject",
                audit_tag="MIN_SCORE",
            )
        )
    return out


def _is_total_row(name: str) -> bool:
    return "合计" in name


def _iter_dimension_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    layout: SheetLayout,
) -> Iterable[tuple[tuple, tuple]]:
    """Yield (score_row, standard_row) pairs starting from layout.data_start_row.

    A "score row" is any row whose dimension-name column is non-empty. The next
    row is treated as its standard row only when layout.row_stride == 2 AND that
    next row's name column is empty (a continuation). Otherwise the standard row
    is a blank tuple. Iteration stops after yielding a row whose dimension-name
    column contains "合计".
    """
    rows = list(ws.iter_rows(min_row=layout.data_start_row, values_only=True))
    width = ws.max_column or 0
    blank = (None,) * width
    i = 0
    while i < len(rows):
        score_row = rows[i]
        name_cell = score_row[1] if len(score_row) > 1 else None
        weight_cell = score_row[2] if len(score_row) > 2 else None
        if name_cell is None and weight_cell is None:
            i += 1
            continue
        # Pair with the next row as the standard row only if it's a continuation
        # (i.e., its name column is empty). This keeps single-row dimensions
        # (e.g., layout A's 学历/年龄/英语 and all of layout B) from absorbing
        # the next real dimension or the totals row.
        nxt = rows[i + 1] if i + 1 < len(rows) else blank
        nxt_name = nxt[1] if len(nxt) > 1 else None
        if layout.row_stride == 2 and nxt_name is None:
            standard_row = nxt
            advance = 2
        else:
            standard_row = blank
            advance = 1
        yield (score_row, standard_row)
        if isinstance(name_cell, str) and _is_total_row(name_cell):
            return
        i += advance


def _scan_trailing_hard_filters(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> list[HardFilter]:
    """Scan every string cell for hard-filter heuristics (age / min score)."""
    out: list[HardFilter] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                out.extend(_parse_hard_filters(cell))
    return out


def _parse_grade_thresholds(
    score_row: tuple,
    standard_row: tuple,
    layout: SheetLayout,
) -> list[GradeThreshold]:
    """Parse grade thresholds from the totals row using the layout's tier columns.

    Each tier cell is expected to contain a numeric range like "40 ~ 54". Cells
    without a range (e.g. layout B's dash placeholders) are skipped silently.
    """
    n = len(layout.tier_cols)
    labels = [f"L{i + 1}" for i in range(n)]
    out: list[GradeThreshold] = []
    range_pat = re.compile(r"(\d+)\s*[~～\-]\s*(\d+)")
    for idx, (grade, col) in enumerate(zip(labels, layout.tier_cols)):
        cell = score_row[col] if col < len(score_row) else None
        if not cell:
            continue
        m = range_pat.search(str(cell))
        if not m:
            continue
        low = float(m.group(1))
        desc_cell = standard_row[col] if standard_row and col < len(standard_row) else None
        label_text = (
            str(desc_cell).split("\n")[0].strip()
            if desc_cell
            else str(cell).split("\n", 1)[-1].strip()
        )
        out.append(GradeThreshold(grade=grade, min=low, label=label_text[:64]))
    out.sort(key=lambda g: g.min)
    return out


def _pick_method(name: str, keywords: list[str]) -> str:
    if "学历" in name:
        return "lookup"
    if "经验" in name or "全流程" in name:
        return "experience_years"
    return "tiered_keyword_match"


def _education_table(name: str) -> dict[str, float] | None:
    if "学历" not in name:
        return None
    return {"本科": 12, "专升本": 9, "大专": 6}


def import_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    jd_code: str,
    layout: SheetLayout,
) -> RuleSchema:
    rule_dims: list[RuleDimension] = []
    judge_dims: list[JudgeDimension] = []
    grade_thresholds: list[GradeThreshold] = []
    hard_filters: list[HardFilter] = []

    high_idx_min = next(
        (i for i, lbl in enumerate(layout.tier_labels) if lbl in ("mid", "high")), 0
    )

    for score_row, standard_row in _iter_dimension_rows(ws, layout):
        name_cell = score_row[1] if len(score_row) > 1 else None
        if not name_cell:
            continue
        name = str(name_cell).strip()

        if _is_total_row(name):
            grade_thresholds = _parse_grade_thresholds(score_row, standard_row, layout)
            continue

        weight = _parse_score(score_row[2]) if len(score_row) > 2 else 0.0
        weight = weight or 0.0

        keyword_cell = (
            score_row[layout.keyword_col]
            if layout.keyword_col < len(score_row)
            else None
        )
        keywords = _split_keywords(
            keyword_cell if isinstance(keyword_cell, str) else None
        )

        tiers: list[Tier] = []
        for idx, (col, label) in enumerate(zip(layout.tier_cols, layout.tier_labels)):
            sc = _parse_score(score_row[col]) if col < len(score_row) else None
            if sc is None:
                continue
            tiers.append(
                Tier(
                    label=label,  # type: ignore[arg-type]
                    score=sc,
                    keywords=keywords if idx >= high_idx_min else [],
                )
            )

        dim_id = re.sub(r"\W+", "_", name).strip("_").lower()[:32]
        if _is_judge_dimension(name):
            judge_dims.append(
                JudgeDimension(
                    id=dim_id or f"j{len(judge_dims)}",
                    name=name,
                    weight=weight,
                    prompt_hint=f"证据：{', '.join(keywords) if keywords else name}",
                    tiers=tiers
                    + [Tier(label="unknown", score=None, note="证据不足建议面试时考察")],
                )
            )
        else:
            rule_dims.append(
                RuleDimension(
                    id=dim_id or f"d{len(rule_dims)}",
                    name=name,
                    weight=weight,
                    method=_pick_method(name, keywords),
                    tiers=tiers,
                    table=_education_table(name),
                )
            )

    # Hard-filter notes live in trailing rows; scan whole sheet to collect them.
    hard_filters.extend(_scan_trailing_hard_filters(ws))

    # Dedup by audit_tag
    seen: set[str] = set()
    unique_filters: list[HardFilter] = []
    for hf in hard_filters:
        if hf.audit_tag in seen:
            continue
        seen.add(hf.audit_tag)
        unique_filters.append(hf)

    # Normalize weights so they sum to 100
    total = sum(d.weight for d in rule_dims) + sum(d.weight for d in judge_dims)
    if total and abs(total - 100) > 0.5:
        factor = 100 / total
        for d in rule_dims:
            d.weight = round(d.weight * factor, 2)
        for d in judge_dims:
            d.weight = round(d.weight * factor, 2)

    return RuleSchema(
        version="v1",
        jd_code=jd_code,
        total_score=100,
        passing_threshold=40,
        hard_filters=unique_filters,
        rule_dimensions=rule_dims,
        judge_dimensions=judge_dims,
        grade_thresholds=grade_thresholds,
    )


def import_workbook(path: Path) -> list[RuleSchema]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[RuleSchema] = []
    for sheet_name, jd_code in JD_CODE_BY_SHEET.items():
        if sheet_name not in wb.sheetnames:
            continue
        layout = SHEET_LAYOUT[sheet_name]
        out.append(import_sheet(wb[sheet_name], jd_code, layout))
    return out
